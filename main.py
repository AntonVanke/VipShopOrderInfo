__version__ = "0.1.1"
__author__ = "AntonVanke"
__email__ = "antonvanke@hotmail.com"
__github__ = "https://github.com/AntonVanke/VipShopOrderInfo"

import os
import io
import time
import tkinter as tk
from tkinter import messagebox

import requests  # 用于请求登录二维码
import openpyxl  # 用于保存数据到 Excel
from PIL import ImageTk, Image  # 显示登录二维码
from tinydb import TinyDB, where, Query  # 数据库
from logzero import logger, logfile  # 日志记录

from viptool import VipShopUser, BrowserWeb

logger.info(f"版本号: {__version__}，Github：{__github__}")
# 初始化 Browser 服务
browser = BrowserWeb()

# 判断是否有 USE_LICENSE 文件
FIRST_START = not os.path.exists("USE_LICENSE")

# 设置 LOGGER 存储位置
logfile("viptool.log", encoding="utf-8", maxBytes=1_000_000, backupCount=5)
logger.warning("第一次启动加载中") if FIRST_START else 0
# 定义窗口大小
WIDTH = HEIGHT = 620

# 加载数据库
db = TinyDB("database.json")
users = db.table("users")
orders = db.table("orders")
details = db.table("details")


def ui_close():
    """
    退出应用程序
    :return:
    """
    if messagebox.askokcancel("退出应用程序", "确定退出应用程序？"):
        logger.info("===============退出应用程序===============")
        root.destroy()


def update_users():
    """
    更新账号信息的同时更新界面
    :return:
    """
    for _user in users.all():
        # 依次判断账号是否有效
        if VipShopUser(uid=_user["uid"], token=_user["token"]).is_visible():
            # 有效的话更新有效时间
            users.update({"status": True, "update_time": int(time.time())}, where("uid") == _user["uid"])
        else:
            # 无效账号，提示更新
            logger.warning(f"账号：{_user['uid']}<{_user['mobile']}>[{_user['remarks']}]已失效")
            users.update({"status": False, "update_time": int(time.time())}, where("uid") == _user["uid"])

    for _user in users.search(where("status") == True):
        # 获取账号信息
        user_info = VipShopUser(uid=_user["uid"], token=_user["token"]).get_user_info()["data"]
        users.update(
            {"mobile": user_info["mobile"], "username": user_info["userName"], "nickname": user_info["nickname"],
             "update_time": int(time.time())},
            where("uid") == _user["uid"])
    # 删除旧数据
    account_listbox.delete(0, tk.END)

    # 筛洗出有效信息
    for user in users.all():
        user_nsi = {k: v for k, v in user.items() if k in ["username", "remarks", "status", "nickname"]}

        account_listbox.insert(tk.END,
                               f'{user_nsi["username"]}-{user_nsi["remarks"]}-{"有效" if user_nsi["status"] else "失效"}-{user_nsi["nickname"]}')
    _info = f"账号总数：{len(users.all())}\t有效账号：{len(users.search(where('status') == True))}"
    logger.debug("{更新账号完成}\t" + _info)
    info_text_label.config(text=_info + f"\n{'点击下方[获取登录二维码]登录账号' if not users.all() else ''}")
    return True


def delete_user():
    """
    删除用户
    :return:
    """
    # 获取选择的用户
    _index = account_listbox.curselection()

    # 判断是否选择了用户
    if _index:
        # 删除用户
        _d_user = users.all()[_index[0]]
        logger.info(f"删除账号：{_d_user['uid']}<{_d_user['mobile']}>[{_d_user['remarks']}]")
        users.remove(where("uid") == _d_user["uid"])

    # 更新用户
    update_users()


def user_select(_):
    # 获取选择的账号
    _index = account_listbox.curselection()
    if _index:
        _i = users.search(where("uid") == users.all()[_index[0]]["uid"])[0]
        logger.info(f"选择账号：{_i['uid']}<{_i['mobile']}>[{_i['remarks']}]")
        # 获取用户的订单数量
        orders_count = len(set([item["order_id"] for item in orders.search(where("user_id") == _i["uid"])]))
        info_text_label.config(
            text=f"UID：{_i['uid']}\n电话：{_i['mobile']}\n当前状态：{'cookie有效' if _i['status'] else '请更新cookie'}\n备注：{_i['remarks']}\n数据库订单数量：{orders_count}\n添加时间：{time.strftime('%Y年%m月%d日%H时%M分', time.localtime(_i['add_time']))}\n更新时间：{time.strftime('%Y年%m月%d日%H时%M分', time.localtime(_i['update_time']))}")

        # 挂载操作按钮
        user_delete_button.pack()


def get_qr_code():
    """
    获取登录的二维码
    :return:
    """
    # 由于函数退出后图片就会消失，所以要有这个
    global tk_pic
    # 获取登录的二维码
    pic = requests.get(browser.get_login_url()).content
    pic = Image.open(io.BytesIO(pic))
    tk_pic = ImageTk.PhotoImage(pic)
    qrcode_label.config(image=tk_pic)
    scan_success_button.pack()
    qrcode_label.pack()

    tools_info_label.config(text="请使用手机 app 扫描下方二维码")


def get_cookie():
    """
    获取登录的 cookie
    :return:
    """
    # 卸载控件
    qrcode_label.pack_forget()
    cookie = browser.get_cookie()
    if not cookie:
        info_text_label.config(text="Cookie 获取失败")
    else:
        if not users.search(where("uid") == cookie[0]):
            # 如果用户不存在
            users.insert({"uid": cookie[0], "token": cookie[1], "status": True, "add_time": int(time.time()),
                          "update_time": int(time.time()), "remarks": "备注"})
            # 如果用户存在
        else:
            users.update({"token": cookie[1], "status": True, "update_time": int(time.time())},
                         where("uid") == cookie[0])

        tools_info_label.config(text="添加/更新成功")
        # 更新用户列表
        update_users()

    scan_success_button.pack_forget()


def update_orders():
    for user in users.search(where("status") == True):
        data = VipShopUser(uid=user["uid"], token=user["token"]).get_orders(page_size=100)
        if data["msg"] == "success":
            for order in data["data"]["orders"]:
                for index_goods in range(len(order["goodsView"])):
                    # 商品编号
                    product_id = order["goodsView"][index_goods]["vSkuId"]
                    # 商品名称
                    product_name = order["goodsView"][index_goods]["name"]
                    # 商品型号
                    product_size = order["goodsView"][index_goods]["sizeName"]
                    # 型号数量
                    product_quantity = order["goodsView"][index_goods]["num"]
                    # 商品价格
                    product_price = order["goodsView"][index_goods]["price"]
                    # 订单号
                    order_id = order["orderSn"]
                    # 用户ID
                    user_id = user["uid"]
                    # 订单状态
                    #   退款完成   97
                    #   已完成     60
                    #   已签收     25
                    #   已发货     22
                    #   订单已取消  97
                    order_status = order["orderStatusName"]
                    # 订单时间
                    # order_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(order["createTime"]))
                    order_time = order["createTime"]
                    # 订单价格
                    order_amount = order["orderAmount"]
                    # 爬取时间
                    # crawl_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
                    crawl_time = int(time.time())

                    # 如果已存在，只修改状态，否则插入
                    order_target = orders.search((where("order_id") == order_id) & (where("product_id") == product_id))
                    if order_target:
                        orders.update({"order_status": order_status},
                                      (where("order_id") == order_id) & (where("product_id") == product_id))
                    else:
                        orders.insert(
                            {"product_id": product_id, "product_name": product_name, "product_size": product_size,
                             "product_quantity": product_quantity, "product_price": product_price, "order_id": order_id,
                             "user_id": user_id, "order_status": order_status, "order_time": order_time,
                             "order_amount": order_amount, "crawl_time": crawl_time})

    return orders.all()


def update_details():
    for user in users.search(where("status") == True):
        orders_sn = list(set([item['order_id'] for item in orders.search(where("user_id") == user["uid"])]))
        check_sn = [item['order_sn'] for item in details.all()]
        for osn in orders_sn:

            # Fixme: 减少请求次数 & 减慢请求速度
            if osn in check_sn:
                if time.time() - details.search(where("order_sn") == osn)[0]["order_time"] > 604800:
                    continue
            time.sleep(0.1)

            data = VipShopUser(uid=user["uid"], token=user["token"]).get_details(osn)
            if data["code"] == 200:
                # 订单编号
                order_sn = data["result"]["orderSn"]
                # 订单时间
                order_time = data["result"]["orderTime"]
                # 订单状态
                order_state = data["result"]["orderTrack"]["result"]["currentState"]
                # 用户ID
                user_id = data["result"]["userId"]
                # 订单商品类数
                goods_count = data["result"]["goodsList"]["result"]["goodsTotalCount"]
                # 订单价格
                total_amount = data["result"]["goodsList"]["result"]["totalAmount"]
                # 物流单号
                tracking_sn = data["result"]["orderTrack"]["result"]["orderTrack"][0]["sn"]
                # 物流信息
                tracking_message = data["result"]["orderSummary"]["result"]["message"]
                # 物流电话
                tracking_mobile = data["result"]["orderSummary"]["result"]["mobile"]
                # 物流收件人
                tracking_consignee = data["result"]["orderSummary"]["result"]["consignee"]
                # 物流地址
                tracking_address = data["result"]["orderSummary"]["result"]["areaName"] + \
                                   data["result"]["orderSummary"]["result"]["detailAddress"]
                # 爬取时间
                crawl_time = int(time.time())

                # 如果已存在，只修改状态，否则插入
                order_target = details.search((where("order_sn") == order_sn))
                if order_target:
                    details.update({"order_state": order_state}, (where("order_sn") == order_sn))
                else:
                    details.insert(
                        {"order_sn": order_sn, "order_time": order_time, "order_state": order_state, "user_id": user_id,
                         "goods_count": goods_count, "total_amount": total_amount, "tracking_sn": tracking_sn,
                         "tracking_message": tracking_message, "tracking_mobile": tracking_mobile,
                         "tracking_consignee": tracking_consignee, "tracking_address": tracking_address,
                         "crawl_time": crawl_time})
    return details.all()


def update_orders_excel(_update=True):
    """
    导出 excel 的详细内容
    :return:
    """
    if _update:
        data = update_orders()
    else:
        data = orders.all()
    if not len(data):
        logger.error("没有订单数据")
        return
    crawl_time = data[-1]['crawl_time']
    workbook = openpyxl.Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = f"截至{time.strftime('%Y-%m-%d %H.%M', time.localtime(data[-1]['crawl_time']))}订单数据"

    headers = ["商品编号", "商品名称", "商品规格", "商品数量", "商品价格", "订单编号", "用户编号", "订单状态",
               "下单时间", "订单总金额", "抓取时间"]

    for i, header in enumerate(headers):
        orders_sheet.cell(row=1, column=i + 1).value = header

    # 写入数据到工作表中，从第二行开始
    for i, item in enumerate(data):
        # 遍历每个字典的键和值
        for j, (key, value) in enumerate(item.items()):
            # 写入工作表的单元格中，根据键名的顺序和值的类型
            if key in ["order_time", "crawl_time"]:
                orders_sheet.cell(row=i + 2, column=j + 1).value = time.strftime('%Y-%m-%d %H:%M',
                                                                                 time.localtime(value))
            elif key in ["product_price", "order_amount"]:
                orders_sheet.cell(row=i + 2, column=j + 1).value = float(value)
            elif key in ["product_quantity"]:
                orders_sheet.cell(row=i + 2, column=j + 1).value = int(value)
            else:
                orders_sheet.cell(row=i + 2, column=j + 1).value = value
    if _update:
        # 创建订单详情数据表
        data = update_details()
        details_sheet = workbook.create_sheet("订单详细数据", 2)
        headers = ["订单编号", "下单时间", "订单状态", "用户编号", "订单商品总数", "订单总金额", "物流单号", "物流信息",
                   "收货人电话", "收货人", "收货地址", "爬取时间"]
        for i, header in enumerate(headers):
            details_sheet.cell(row=1, column=i + 1).value = header

        # 写入数据到工作表中，从第二行开始
        for i, item in enumerate(data):
            # 遍历每个字典的键和值
            for j, (key, value) in enumerate(item.items()):
                # 写入工作表的单元格中，根据键名的顺序和值的类型
                if key in ["order_time", "crawl_time"]:
                    details_sheet.cell(row=i + 2, column=j + 1).value = time.strftime('%Y-%m-%d %H:%M',
                                                                                      time.localtime(value))
                elif key in ["total_amount"]:
                    details_sheet.cell(row=i + 2, column=j + 1).value = float(value)
                elif key in ["goods_count"]:
                    details_sheet.cell(row=i + 2, column=j + 1).value = int(value)
                else:
                    details_sheet.cell(row=i + 2, column=j + 1).value = value

    # 创建用户表
    data = users.all()
    users_sheet = workbook.create_sheet("当前用户状态", 2)
    headers = ["用户编号", "TOKEN", "状态", "添加时间", "更新时间", "备注", "电话", "用户名", "用户昵称"]
    for i, header in enumerate(headers):
        users_sheet.cell(row=1, column=i + 1).value = header

    for i, item in enumerate(data):
        # 遍历每个字典的键和值
        for j, (key, value) in enumerate(item.items()):
            # 写入工作表的单元格中，根据键名的顺序和值的类型
            if key in ["update_time", "add_time"]:
                users_sheet.cell(row=i + 2, column=j + 1).value = time.strftime('%Y-%m-%d %H:%M',
                                                                                time.localtime(value))
            elif key == "status":
                users_sheet.cell(row=i + 2, column=j + 1).value = "有效账号" if value else "无效账号"
            else:
                users_sheet.cell(row=i + 2, column=j + 1).value = value

    # 保存工作簿到文件中
    try:
        workbook.save(f"orders-{time.strftime('%Y年%m月%d日%H时%M分', time.localtime(crawl_time))}.xlsx")
    except Exception as err:
        logger.exception(err)
    tools_info_label.config(
        text=f"导出文件：orders-{time.strftime('%Y年%m月%d日%H时%M分', time.localtime(crawl_time))}.xlsx")
    logger.info(f"导出文件成功：orders-{time.strftime('%Y年%m月%d日%H时%M分', time.localtime(crawl_time))}.xlsx")


if __name__ == '__main__':
    logger.info("---------------启动应用程序---------------")
    root = tk.Tk()

    # 设置页面基本信息
    root.title("唯品会订单导出工具")
    root.geometry(f"{WIDTH}x{HEIGHT}")
    # 载入应用图标
    try:
        root.iconbitmap("favicon.ico")
    except tk.TclError as err:
        logger.exception(err)

    # 第一次启动弹出协议
    if FIRST_START:
        license_text = f"1. 本程序为自用开源程序\n2. 使用风险需自行承担\n3. 使用后账号出现砍单情况请反馈作者<{__email__}>\n请点击确认同意协议"
        if messagebox.askokcancel("使用协议", license_text):
            logger.info("*****您已同意协议*****")
            with open("USE_LICENSE", "w", encoding="utf-8") as _license:
                _license.write(license_text + "\n\n\t我已同意以上全部内容")
        else:
            root.destroy()

    # 组件
    account_labelframe = tk.LabelFrame(root, text="账号列表")
    account_listbox = tk.Listbox(account_labelframe)

    account_op_labelframe = tk.LabelFrame(root, text="账号信息")
    info_text_label = tk.Label(account_op_labelframe, text=f"账号总数：{len(users.all())}\t"
                                                           f"有效账号：{len(users.search(where('status') == True))}")
    update_users_button = tk.Button(account_op_labelframe, text="更新所有账号", command=update_users)
    user_delete_button = tk.Button(account_op_labelframe, text="删除账号", command=delete_user)

    tools_labelframe = tk.LabelFrame(root, text="操作")
    tools_info_label = tk.Label(tools_labelframe, text="点击 (获取登录二维码) 后使用手机APP扫码")
    qrcode_label = tk.Label(tools_labelframe, background="#ffffff")
    get_qrcode_button = tk.Button(tools_labelframe, text="获取登录二维码", command=get_qr_code)
    scan_success_button = tk.Button(tools_labelframe, text="我已扫码登录成功", command=get_cookie)
    update_express_button = tk.Button(tools_labelframe, text="更新所有物流数据", command=update_orders, fg="#a27af4")
    update_all_button = tk.Button(tools_labelframe, text="更新所有订单数据", command=update_orders, fg="#2a42b4")
    export_data_button = tk.Button(tools_labelframe, text="仅导出数据库订单数据",
                                   command=lambda: update_orders_excel(False), fg="#599e5e")
    export_all_data_button = tk.Button(tools_labelframe, text="更新所有订单/物流数据并导出订单",
                                       command=update_orders_excel,
                                       fg="#c94f4f")

    # 挂载
    account_labelframe.place(x=10, y=10, width=300, height=HEIGHT - 2 * 10)
    account_op_labelframe.place(x=315, y=10, width=300, height=210)
    tools_labelframe.place(x=315, y=220, width=300, height=390)
    account_listbox.place(x=5, y=5, width=285, height=HEIGHT - 6 * 10)

    info_text_label.pack()
    tools_info_label.pack()
    update_users_button.pack()
    get_qrcode_button.pack()
    update_all_button.pack()
    # update_express_button.pack()
    export_data_button.pack()
    export_all_data_button.pack()

    update_users()
    # 绑定账号列表选择
    account_listbox.bind("<<ListboxSelect>>", func=user_select)
    # 捕获关闭命令
    root.protocol('WM_DELETE_WINDOW', ui_close)
    # 保持窗口
    root.mainloop()
